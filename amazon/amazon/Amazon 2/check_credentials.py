import openpyxl
from openpyxl import load_workbook

path = 'C:\\Users\\Kunal Talwar\\Desktop\\Python\\test_login.xlsx'


def check_test_readdata():
    list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_rows
    cols = sheet.max_columns
    print(rows)
    print(cols)
    for r in range(2, rows + 1):
        # for c in range(1, cols+1):
        # sheet.cell(r, c).value

        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple = (username, password)
        list.append(tuple)


def readData(file,rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    # print(sheet.cell(row=rownum, column=columnno).value)
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetname,rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value =data
    workbook.save(file)

# readData(path,2,2)



'''
workbook = openpyxl.load_workbook('C:\\Users\\Kunal Talwar\\Desktop\\Python\\test_login.xlsx')
sheet = workbook.active

print(sheet.cell(row=2,column=2).value)

'''