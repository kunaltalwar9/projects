import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl
import pytest

driver = webdriver.Firefox()

driver.get("https://fasalrin.gov.in/")
# Click on Users Dropdown
driver.find_element(By.XPATH,"//div[@class='left-icon bottom-left']/div[1]").click()
# Click on Login
driver.find_element(By.XPATH, "//div[@class='left-icon bottom-left']/div[2]" ).click()
time.sleep((0.5))
# Enter Mobile Number
driver.find_element(By.XPATH, "//input[@name='username']").send_keys("8950080781")
# Enter Password
driver.find_element(By.XPATH,"//input[@name='loginPwd']").send_keys("Shgb@8043")
# Wait time to Enter Captcha Manually
time.sleep(9)
# Login Button Click
driver.find_element(By.CSS_SELECTOR, ".genGreenBtn > span:nth-child(1)").click()
wait = WebDriverWait(driver, 600)
wait.until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"a[title='Loan Application']")))
# Popup Accept
driver.find_element(By.CSS_SELECTOR,"button.btn:nth-child(1) > span:nth-child(1)").click()
# Click on Loan Application
driver.find_element(By.CSS_SELECTOR,"a[title='Loan Application']").click()
# Select Financial year
select1 = Select(driver.find_element(By.CSS_SELECTOR,"select[name='financialYear']")).select_by_index(1)
# Adding Aadhar Number via excel
path = "C:\\Users\\Kunal Talwar\\Downloads\\2023-24 Rin ac data.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]
# Starting from Row 2
'''
for i in range(3,6):
    # Aadhar Record Not Added
    if sheet.cell(row=i,column=2).value == 'Not Present':
        print("Not Present")
        aadhar_number = sheet.cell(row=i,column=3).value
        print(aadhar_number)
        # Enter Aadhar Number
        driver.find_element(By.CSS_SELECTOR,"div.modal-body:nth-child(1) > div:nth-child(1) > div:nth-child(2) > form:nth-child(1) > div:nth-child(1) > input").send_keys(aadhar_number)
        # Submit Aadhar and FY
        driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn mt-3 green-btn aadharBtnWidth']").click()
        time.sleep(1)
        # Select Account Number
        select2= Select(driver.find_element(By.CSS_SELECTOR,"select[name='accountNumbers']")).select_by_index(1)
        # Click Ok
        driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-btn mt-3 ms-md-3 green-btn']").click()
        # Select Application type
        select3 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='applicationType']")).select_by_index(1)
        # Select Village
        driver.execute_script("window.scrollBy(0,20000)", "")
        time.sleep(1)
'''
for i in range(2,137):
    # Aadhar Record Not Added
    aadhar_number = sheet.cell(row=i,column=2).value
    print(aadhar_number)
    # Enter Aadhar Number
    driver.find_element(By.CSS_SELECTOR,"div.modal-body:nth-child(1) > div:nth-child(1) > div:nth-child(2) > form:nth-child(1) > div:nth-child(1) > input").send_keys(aadhar_number)
    # Submit Aadhar and FY
    driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn mt-3 green-btn aadharBtnWidth']").click()
    time.sleep(1)
    # Select Account Number
    # select2= Select(driver.find_element(By.CSS_SELECTOR,"select[name='accountNumbers']")).select_by_index(1)
    # Click Ok
    driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-btn mt-3 ms-md-3 green-btn']").click()
    # Select Application type
    select3 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='applicationType']")).select_by_index(1)
    # Type Name
    driver.find_element(By.CSS_SELECTOR, "input[name='beneficiaryName']").send_keys(sheet.cell(row=i,column=3).value)
    # Click on Verify
    driver.find_element(By.CSS_SELECTOR,"button[class='btn btn-btn purple-btn applicant_input_inner_btn']").click()
    time.sleep(10)
    # Type Passbook Name
    driver.find_element(By.CSS_SELECTOR, "input[name='beneficiaryPassbookName']").send_keys(sheet.cell(row=i, column=3).value)
    # Enter DOB
    driver.find_element(By.CSS_SELECTOR, ".inputDateValidation > div:nth-child(1) > div:nth-child(2) > div:nth-child("
                                         "1) > div:nth-child(1) > input:nth-child(1)").send_keys((sheet.cell(row=i,
                                                                                                             column=6).value).strftime('%d/%m/%Y'))
    time.sleep(10)
    # Select Gender
    if sheet.cell(row=i, column=14).value == 'Male':
        select4 = Select(driver.find_element(By.CSS_SELECTOR, "select[id='gender']")).select_by_index(1)
    elif sheet.cell(row=i, column=14).value == 'Female':
        select4 = Select(driver.find_element(By.CSS_SELECTOR, "select[id='gender']")).select_by_index(2)
    # Enter Phone Number
    driver.find_element(By.CSS_SELECTOR, "input[name='mobile']").send_keys(sheet.cell(row=i, column=5).value)
    # Select Social Category
    select5 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='casteCategory']")).select_by_index(3)
    # Select Farmer Category
    select6 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='farmerCategory']")).select_by_index(1)
    # Select Farmer Type
    driver.execute_script("window.scrollBy(0,20000)", "")
    time.sleep(1)
    select7 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='farmerType']")).select_by_index(3)
    # Select Relative type
    if sheet.cell(row=i, column=14).value == 'Male':
        select8 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='relation']")).select_by_index(1)
    elif sheet.cell(row=i, column=14).value == 'Female':
        select8 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='relation']")).select_by_index(2)
    # Type Father Name
    driver.find_element(By.CSS_SELECTOR, "input[name='relativeName']").send_keys(sheet.cell(row=i, column=4).value)
    # Select Block/District
    select9 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='resSubDistrictId']")).select_by_index(2)
    # Select Village
    '''
        print(str(sheet.cell(row=i,column=8).value))
        if str(sheet.cell(row=i,column=8).value)=="Phulpura (3) (Phoolpura)":
            print("Pass")
        else:
            print("Fail")
    '''
    # element = driver.find_element(By.CSS_SELECTOR, "select[name='resVillageId']")
    # driver.execute_script("arguments[0].scrollIntoView();", element)
    print(str(sheet.cell(row=i,column=7).value))
    assert str(sheet.cell(row=i,column=7).value) == "Kharak Kalan (131) (Kharak Kalan)"
    driver.execute_script("window.scrollBy(0,20000)", "")
    select10 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='resVillageId']")).select_by_visible_text(str(sheet.cell(row=i,column=7).value))#select_by_index(2)
    # Clear Address
    driver.find_element(By.CSS_SELECTOR, "input[name='resAddress']").clear()
    # Add Address
    driver.find_element(By.CSS_SELECTOR, "input[name='resAddress']").send_keys(str(sheet.cell(row=i,column=7).value))
    # Enter Pincode
    driver.find_element(By.CSS_SELECTOR, "input[name='resPincode']").send_keys("127114")
    # Select Continue
    driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn green-btn']").click()
    time.sleep(1)
    # Select Continue
    driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn green-btn ms-auto']").click()
    # Sanction Amount Add
    driver.find_element(By.CSS_SELECTOR, "input[name='loanSanctionAmount']").send_keys(str(sheet.cell(row=i,column=5).value))
    # Drawing Limit Add
    driver.find_element(By.CSS_SELECTOR, "input[name='drawingLimit']").send_keys(str(sheet.cell(row=i, column=6).value))
    # Date Input
    time.sleep(8)
    # driver.find_element(By.CSS_SELECTOR, "#finance > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > input:nth-child(1)").send_keys("31/03/2023")
    # driver.find_element(By.CSS_SELECTOR,"#finance > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > input:nth-child(1)").click()
    # back_button_on_calendar= driver.find_element(By.CSS_SELECTOR, "span[class='rmdp-arrow-container rmdp-left ']")
    # next_button_on_calendar= driver.find_element(By.CSS_SELECTOR, "span[class='rmdp-arrow-container rmdp-right ']")
    # Select Continue
    driver.find_element(By.CSS_SELECTOR, "#finance > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > button:nth-child(2)").click()
    # Select Agri Crops and Horti Crops
    driver.find_element(By.CSS_SELECTOR, "button.act_tab_btn:nth-child(1)").click()
    driver.find_element(By.CSS_SELECTOR, "button.act_tab_btn:nth-child(2)").click()
    # Logout
    # driver.find_element(By.CSS_SELECTOR,"aside.bg-light:nth-child(1) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(7) > a:nth-child(1) > span:nth-child(2)").click()
    # driver.quit()
    # break

    '''
        print(str(sheet.cell(row=i,column=8).value))
        if str(sheet.cell(row=i,column=8).value)=="Phulpura (3) (Phoolpura)":
            print("Pass")
        else:
            print("Fail")
    '''
    '''
        # element = driver.find_element(By.CSS_SELECTOR, "select[name='resVillageId']")
        # driver.execute_script("arguments[0].scrollIntoView();", element)
        select4 = Select(driver.find_element(By.CSS_SELECTOR, "select[name='resVillageId']")).select_by_visible_text(str(sheet.cell(row=i,column=8).value))
        # Clear Address
        driver.find_element(By.CSS_SELECTOR, "input[name='resAddress']").clear()
        # Add Address
        driver.find_element(By.CSS_SELECTOR, "input[name='resAddress']").send_keys(str(sheet.cell(row=i,column=8).value))
        # Select Continue
        driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn green-btn']").click()
        time.sleep(1)
        # Select Continue
        driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-btn green-btn ms-auto']").click()
        # Sanction Amount Add
        driver.find_element(By.CSS_SELECTOR, "input[name='loanSanctionAmount']").send_keys(str(sheet.cell(row=i,column=5).value))
        # Drawing Limit Add
        driver.find_element(By.CSS_SELECTOR, "input[name='drawingLimit']").send_keys(str(sheet.cell(row=i, column=6).value))
        # Date Input
        time.sleep(8)
        # driver.find_element(By.CSS_SELECTOR, "#finance > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > input:nth-child(1)").send_keys("31/03/2023")
        # driver.find_element(By.CSS_SELECTOR,"#finance > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > input:nth-child(1)").click()
        # back_button_on_calendar= driver.find_element(By.CSS_SELECTOR, "span[class='rmdp-arrow-container rmdp-left ']")
        # next_button_on_calendar= driver.find_element(By.CSS_SELECTOR, "span[class='rmdp-arrow-container rmdp-right ']")
        # Select Continue
        driver.find_element(By.CSS_SELECTOR, "#finance > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > button:nth-child(2)").click()
        # Select Agri Crops and Horti Crops
        driver.find_element(By.CSS_SELECTOR, "button.act_tab_btn:nth-child(1)").click()
        driver.find_element(By.CSS_SELECTOR, "button.act_tab_btn:nth-child(2)").click()
        # Logout
        # driver.find_element(By.CSS_SELECTOR,"aside.bg-light:nth-child(1) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(7) > a:nth-child(1) > span:nth-child(2)").click()
        # driver.quit()
        break

    else:
        # Aadhar Record already Added
        print("Record Already Added")
'''
''' if sheet['B133'].value== 'Not Present':
    print("Not Present")
    aadhar_number = sheet['D133'].value
else:
    print("Present")
'''

# Adding Aadhar Number Manually
# driver.find_element(By.CSS_SELECTOR, "div.modal-body:nth-child(1) > div:nth-child(1) > div:nth-child(2) > form:nth-child(1) > div:nth-child(1) > input").send_keys("677478127904")


