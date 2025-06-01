import pandas as pd
import pywhatkit as kit
import time
import openpyxl

# Load the Excel file
file_path = "C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Contacts.xlsx"  # Replace with your file path
# data = pd.read_excel(file_path)
# print("Column names in the Excel file:", data.columns.tolist())
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Sector']
image_path = ("C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Banner.jp "
              ""
              "")


# Ensure your Excel has columns 'Phone' and 'Message'
'''for index, row in data.iterrows():
    phone = str(row['Phone']).strip()
    message = row['Message']
'''
'''
for col in range(11, 12):  # Columns 5 to 10 (inclusive)
    print(f"Starting messages for Column {col}...")
    for row in range(1080, 1195):  # Assuming row 1 is a header
        try:
            # Read phone number and message
            phone = sheet.cell(row=row, column=col).value
            name = sheet.cell(row=row, column=1).value


            # Ensure phone number is a valid string
            if phone is None:
                print(f"Row {row}: Missing phone number. Skipping...")
                continue

            # if not phone.startswith('+'):  # Add country code if missing
            #    phone = '+91' + str(phone).strip()  # Replace '+91' with your country's code
            phone = "+91" + str(phone).strip()  # Add countr


            # Use the constant message
            message = """  
            सभी रोहतक वासियों को सूचित किया जाता है कि जो भी नागरिक अपने वोटर कार्ड में कोई बदलाव करवाना चाहते हैं, जैसे नाम, पता, फोटो या अन्य जानकारी, वे 25 जनवरी, शनिवार को अपने-अपने बूथ पर जाकर इसे अपडेट करवा सकते हैं।

कृपया अपने साथ आवश्यक दस्तावेज़, जैसे वोटर कार्ड, आधार कार्ड, 10वीं की मार्कशीट, बिजली/पानी का बिल और पासपोर्ट साइज फोटो अवश्य लेकर जाएं।

यह सूचना अधिक से अधिक लोगों तक पहुंचाएं। 

धन्यवाद।
कुनाल तलवाड पुत्र टेकचंद तलवाड
पार्षद प्रत्याशी
नगर निगम वार्ड नंबर 12, रोहतक """
            # message = "Namaskar " + str(name)+" Please support Kunal"

            # Debugging: Print the details being sent
            print(f"Sending to: {phone}, Message: {message}, Image: {image_path}")

            # Send WhatsApp message with image
            kit.sendwhats_image(receiver=phone, img_path=image_path, caption=message, wait_time=140, tab_close=True)
            print(f"Message and image sent to {phone}")
            time.sleep(5)  # Pause to avoid rate limits

        except Exception as e:
            print(f"Failed to send message to {phone if 'phone' in locals() else 'Unknown'}: {e}")
            
'''

for row in range(173, 175):  # Assuming row 1 is a header
    try:
        # Read phone number and message
        phone = sheet.cell(row=row, column=6).value
        # phone = '+91' + str(phone).strip()  # Replace '+91' with your country's code
        phone = "+91" + str(phone).strip()  # Add country

        # Use the constant message
        message = """  
        Dear Friends,

With great joy and excitement, I invite you to the inauguration of my election office. Your presence and support will mean a lot to me as I embark on this important journey.

📅 Date: 11th February 2025 (Sunday)
⏰ Time: 11:00 AM
📍 Venue: Shop No. 88-89, Sector 1, Near Vita Booth, HUDA, Rohtak

I look forward to welcoming you and sharing this special moment together. Your blessings and encouragement will be my biggest strength. See you there!

With warm regards,
Kunal Talwar

📞 9896767145 """

        # Send WhatsApp message with image
        kit.sendwhats_image(receiver=phone, img_path=image_path, caption=message, wait_time=140, tab_close=True)
        print(f"Message and image sent to {phone}")
        time.sleep(5)  # Pause to avoid rate limits

    except Exception as e:
        print(f"Failed to send message to {phone if 'phone' in locals() else 'Unknown'}: {e}")