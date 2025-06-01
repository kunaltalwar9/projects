import time
from selenium import webdriver
import openpyxl
import pywhatkit as kit
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
'''
driver = webdriver.Firefox()
driver.get("https://web.whatsapp.com/")
time.sleep(20)

path = "C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Phone Numbers.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

for i in range(1,2):
    phone_number = sheet.cell(row=i,column=5).value
    print(phone_number)
    driver.get(phone_number)
    # time.sleep(1)
    driver.back()
    driver.refresh()
    driver.switch_to.window(driver.current_window_handle)
    # driver.get("https://api.whatsapp.com/send/?phone=%2B919896767145&text&type=phone_number&app_absent=0")
    driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[2]/div/section/div/div/div/div[2]/div[1]/h2/p").click()
    driver.find_element(By.CSS_SELECTOR, "a[id='action-button']").is_displayed()
    driver.find_element(By.CSS_SELECTOR, "a[id='action-button']").is_enabled()
    print("Continue to Chat")


phone_number = sheet.cell(row=1,column=5).value
print(phone_number)
driver.get(phone_number)
'''
'''
driver = webdriver.Firefox()
driver.get("https://web.whatsapp.com/")
time.sleep(50)
path = "C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Phone Numbers.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
phone_number = sheet.cell(row=1,column=5).value
print(phone_number)
driver.get(phone_number)
# driver.get("https://api.whatsapp.com/send/?phone=%2B919896767145&text&type=phone_number&app_absent=0")
# driver.find_element(By.CSS_SELECTOR, "a[id='action-button']").click()
print("Test")
# time.sleep(20)

'''
'''
driver = webdriver.Firefox()
file_path = "C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Contacts.xlsx"  # Replace with your file path
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Whatsapp']
driver.get("https://web.whatsapp.com")
time.sleep(100)

for i in range(1080,1195):
    phone_number = sheet.cell(row=i,column=11).value
    print(phone_number)
    driver.get(f"https://wa.me/+91"+str(phone_number))
    # time.sleep(150000)
'''

#############################################
file_path = "C:\\Users\\Kunal Talwar\\Downloads\\Elektron\\Contacts.xlsx"  # Replace with your file path
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Whatsapp']

for row in range(263, 1033):  # Assuming row 1 is a header
    try:
        # Read phone number and message
        phone = sheet.cell(row=row, column=11).value
        # if not phone.startswith('+'):  # Add country code if missing
        phone = "+91" + str(phone).strip()  # Add country code
        # Use the constant message
        message = """ 
        सभी रोहतक वासियों को सूचित किया जाता है कि जो भी नागरिक अपने वोटर कार्ड में कोई बदलाव करवाना चाहते हैं, जैसे नाम, पता, फोटो या अन्य जानकारी, वे 25 जनवरी, शनिवार को अपने-अपने बूथ पर जाकर इसे अपडेट करवा सकते हैं।

कृपया अपने साथ आवश्यक दस्तावेज़, जैसे वोटर कार्ड, आधार कार्ड, 10वीं की मार्कशीट, बिजली/पानी का बिल और पासपोर्ट साइज फोटो अवश्य लेकर जाएं।

यह सूचना अधिक से अधिक लोगों तक पहुंचाएं। धन्यवाद।

कुनाल तलवाड पुत्र टेकचंद तलवाड
वार्ड नंबर 12, रोहतक

This is to inform all residents of Rohtak that anyone who wishes to make any changes to their voter card, such as name, address, photo, or other details, can visit their respective booth on Saturday, 25th January, to get it updated.

Please ensure you carry the required documents, such as your voter card, Aadhaar card, 10th marksheet, electricity/water bill, and a passport-size photo.

Kindly share this information with as many people as possible. Thank you.

Kunal Talwar
S/O Tekchand Talwar
Ward No. 12, Rohtak
         """


            # Debugging: Print the details being sent
        print("Sending to: {phone}")

        # Send WhatsApp message with image
        kit.sendwhatmsg_instantly(phone_no=phone, message=message, wait_time=25, tab_close=True)
        print(f"Message sent to {phone}")
        # time.sleep(1)  # Pause to avoid rate limits

    except Exception as e:
        print(f"Failed to send message to {phone if 'phone' in locals() else 'Unknown'}: {e}")