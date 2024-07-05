import openpyxl
import pytest
import allure
import time

from openpyxl.compat import file
from selenium import webdriver


@pytest.fixture()
def test_setup():
    global driver
    driver = webdriver.Chrome(
        executable_path=r"C:\Users\Kunal Talwar\eclipse-workspace\libs\chromedriver_win32\chromedriver.exe")
    time.sleep(3)
    driver.maximize_window()
    yield
    driver.quit()


def test_readdata():
    list = []
    path = "C:\\Users\\Kunal Talwar\\Desktop\\test_Testdata.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_rows
    cols = sheet.max_columns
    print(rows)
    print(cols)
    for r in range(2, rows + 1):
        # for c in range(1, cols+1):
        # sheet.cell(r, c).value

        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple = (username, password)
        list.append(tuple)


@allure
def test_signin():
    driver.get("https://www.zales.com")
    print(driver.current_url)
    driver.maximize_window()
    driver.find_element_by_css_selector(
        "#main-container > div.main-inner > header > div.top-navigation.hidden-xs.hidden-sm > div > div.col-md-7.my-account-navigation.hidden-sm.hidden-xs.pull-right > ul > li.top-nav-link.sign-in > a.sidebar-login").click()
    # driver.find_element_by_xpath("/html/body/main/div/div[2]/div[3]/div[6]/div[1]/nav/ul/li/div/div[1]/div[1]/ul/li[1]/a").click()
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="sign-in"]').click()


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file, sheetname,rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file, sheetname,rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value =data
    workbook.save(file)

#abc = driver.find_element_by_link_text().text
#assert "abc" in abc
#assert "abc" == abc
#implicit wait checks whether element exists, doesnt check if element is visible